"""
api/routes/jobs.py — Async job tracking for long-running bulk operations.

POST /api/jobs/bulk-export   — kick off an async XLSX export; returns job_id
GET  /api/jobs/{job_id}      — poll status: pending | running | complete | failed

For the hackathon, jobs run in a background threading.Thread.
In production this would be Celery + Redis.
"""

from __future__ import annotations

import io
import threading
import time
import uuid
from datetime import date
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse

router = APIRouter(prefix="/api/jobs", tags=["jobs"])

# In-memory job registry: job_id -> status dict
_JOBS: dict[str, dict] = {}
_FILES_DIR = Path(__file__).parent.parent.parent / "tmp_exports"
_FILES_DIR.mkdir(exist_ok=True)


def _get_store():
    from api.main import app
    return app.state.store


def _run_xlsx_export(job_id: str) -> None:
    """Background thread: build XLSX and write to disk, then mark job complete."""
    try:
        _JOBS[job_id]["status"] = "running"
        _JOBS[job_id]["progress"] = 10

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        store = _get_store()
        entries = sorted(store.values(), key=lambda e: e["scored"].risk_score, reverse=True)
        today = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Risk"

        headers = [
            "Vendor ID", "Name", "Category", "Risk Score", "Risk Level",
            "SOC2", "ISO27001", "GDPR DPA", "Contract End",
            "Under Investigation", "Breach Count", "Top Risk Factor",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2D3748")
            cell.alignment = Alignment(horizontal="center")

        _JOBS[job_id]["progress"] = 40

        LEVEL_FILL = {"CRITICAL": "C53030", "HIGH": "C05621", "MEDIUM": "975A16", "LOW": "276749"}
        for row_idx, e in enumerate(entries, 2):
            v = e["vendor"]
            sv = e["scored"]
            level = sv.risk_level.value
            row_data = [
                v.vendor_id, v.name, v.category, sv.risk_score, level,
                v.compliance.soc2_type2, v.compliance.iso27001, v.compliance.gdpr_dpa,
                v.contract_end.isoformat(), v.under_investigation, len(v.breach_history),
                sv.risk_factors[0][:80] if sv.risk_factors else "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 5 and level in LEVEL_FILL:
                    cell.fill = PatternFill("solid", fgColor=LEVEL_FILL[level])
                    cell.font = Font(color="FFFFFF", bold=True)

        _JOBS[job_id]["progress"] = 80

        filename = f"bulk_export_{today.isoformat()}_{job_id}.xlsx"
        filepath = _FILES_DIR / filename
        wb.save(str(filepath))

        _JOBS[job_id]["status"] = "complete"
        _JOBS[job_id]["progress"] = 100
        _JOBS[job_id]["result_url"] = f"/api/jobs/{job_id}/download"
        _JOBS[job_id]["filename"] = filename

    except Exception as exc:
        _JOBS[job_id]["status"] = "failed"
        _JOBS[job_id]["error"] = str(exc)


@router.post("/bulk-export", summary="Kick off async XLSX export job")
def start_bulk_export():
    """Start an async XLSX export. Returns a job_id to poll with GET /api/jobs/{id}."""
    job_id = str(uuid.uuid4())[:8]
    _JOBS[job_id] = {
        "job_id": job_id,
        "status": "pending",
        "progress": 0,
        "result_url": None,
        "error": None,
    }
    t = threading.Thread(target=_run_xlsx_export, args=(job_id,), daemon=True)
    t.start()
    return {"job_id": job_id, "status": "pending"}


@router.get("/{job_id}", summary="Poll async job status")
def get_job_status(job_id: str):
    """Poll a job started via POST /api/jobs/bulk-export."""
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id!r} not found")
    return job


@router.get("/{job_id}/download", summary="Download completed XLSX export")
def download_job_result(job_id: str):
    """Download the XLSX file produced by a completed bulk-export job."""
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id!r} not found")
    if job["status"] != "complete":
        raise HTTPException(status_code=409, detail=f"Job status is {job['status']!r}, not complete")

    filepath = _FILES_DIR / job["filename"]
    if not filepath.exists():
        raise HTTPException(status_code=410, detail="Export file no longer available")

    return FileResponse(
        path=str(filepath),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job["filename"],
    )
