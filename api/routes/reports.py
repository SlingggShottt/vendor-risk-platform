"""
api/routes/reports.py — Portfolio risk report and CSV export endpoints.

Report format mirrors the brief's example:
  - Risk summary (counts per level)
  - Red-flag vendors (CRITICAL + HIGH)
  - Compliance stats (cert coverage, orphaned access, etc.)
  - Actionable recommendations

CSV export returns all scored vendors as a flat CSV.
"""

from __future__ import annotations

import csv
import io
from datetime import date

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response, StreamingResponse

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from common.schema import RiskLevel, ScoredVendor, Vendor
from data.compliance_export import build_compliance_summary, export_to_csv as _build_compliance_csv, export_to_json as _build_compliance_json, format_vendor_for_compliance_export

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _get_store():
    from api.main import app
    return app.state.store


@router.get("")
def portfolio_report():
    """
    Full portfolio risk report. Matches the brief's report format:
    risk summary, red-flag vendors, compliance stats.
    """
    store = _get_store()
    entries = list(store.values())
    today = date.today()

    vendors:  list[Vendor]       = [e["vendor"]  for e in entries]
    scored:   list[ScoredVendor] = [e["scored"]  for e in entries]

    # ── Risk level counts ─────────────────────────────────────────────────────
    level_counts = {lvl.value: 0 for lvl in RiskLevel}
    for sv in scored:
        level_counts[sv.risk_level.value] += 1

    # ── Compliance stats ──────────────────────────────────────────────────────
    total = len(vendors)
    n_soc2_ok  = sum(1 for v in vendors
                     if v.compliance.soc2_type2
                     and (v.compliance.soc2_expiry is None or v.compliance.soc2_expiry >= today))
    n_iso      = sum(1 for v in vendors if v.compliance.iso27001)
    n_gdpr_ok  = sum(1 for v in vendors if not v.handles_eu_data or v.compliance.gdpr_dpa)
    n_soc2_expiring = sum(
        1 for v in vendors
        if v.compliance.soc2_type2
        and v.compliance.soc2_expiry
        and 0 <= (v.compliance.soc2_expiry - today).days <= 60
    )
    n_orphaned = sum(
        1 for v in vendors
        if v.contract_end < today and bool(v.data_access.systems)
    )
    n_under_inv = sum(1 for v in vendors if v.under_investigation)
    n_breached_recent = sum(
        1 for v in vendors
        if v.breach_history
        and (today - max(b.date for b in v.breach_history)).days / 30.44 <= 12
    )

    # ── Red-flag vendors (CRITICAL + HIGH) ────────────────────────────────────
    red_flags = [
        {
            "vendor_id": e["vendor"].vendor_id,
            "name": e["vendor"].name,
            "category": e["vendor"].category,
            "risk_level": e["scored"].risk_level.value,
            "risk_score": e["scored"].risk_score,
            "anomaly_type": e["scored"].anomaly_type.value,
            "top_risk_factor": e["scored"].risk_factors[0] if e["scored"].risk_factors else "",
            "recommendation": e["scored"].recommendation,
        }
        for e in entries
        if e["scored"].risk_level in (RiskLevel.CRITICAL, RiskLevel.HIGH)
    ]
    red_flags.sort(key=lambda x: x["risk_score"], reverse=True)

    # ── Alert counts ──────────────────────────────────────────────────────────
    all_alerts = [a for e in entries for a in e.get("alerts", [])]
    alert_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for a in all_alerts:
        alert_counts[a.severity] = alert_counts.get(a.severity, 0) + 1

    return {
        "generated_at": today.isoformat(),
        "total_vendors": total,
        "risk_summary": level_counts,
        "compliance_stats": {
            "soc2_coverage_pct": round(100 * n_soc2_ok / total, 1) if total else 0,
            "iso27001_coverage_pct": round(100 * n_iso / total, 1) if total else 0,
            "gdpr_compliance_pct": round(100 * n_gdpr_ok / total, 1) if total else 0,
            "soc2_expiring_60d": n_soc2_expiring,
            "orphaned_access_count": n_orphaned,
            "under_investigation_count": n_under_inv,
            "recently_breached_count": n_breached_recent,
        },
        "alert_summary": alert_counts,
        "red_flag_vendors": red_flags,
        "top_recommendations": [
            "Immediately review all CRITICAL vendors with the CISO.",
            "Revoke access for vendors with expired contracts (orphaned access).",
            f"Renew SOC 2 certifications for {n_soc2_expiring} vendor(s) expiring within 60 days.",
            "Require GDPR Data Processing Agreements from all EU-data vendors.",
        ],
    }


@router.post("/email")
def send_email_report():
    """Send the monthly portfolio risk summary email via the configured backend."""
    store = _get_store()
    entries = list(store.values())
    if not entries:
        raise HTTPException(status_code=422, detail="No vendors in portfolio — nothing to report.")

    scored: list[ScoredVendor] = [e["scored"] for e in entries]
    today = date.today()

    try:
        from monitoring.emailer import get_emailer
        get_emailer().send_monthly_summary(scored, today)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Email send failed: {exc}") from exc

    try:
        from api.main import app
        app.state.audit_logger.log_event(
            actor="system",
            action="email_report_sent",
            resource_type="report",
            resource_id="monthly_summary",
            old_state={},
            new_state={"vendor_count": len(scored), "date": today.isoformat()},
            reason="manual send via dashboard",
        )
    except Exception:
        pass

    return {"status": "sent", "vendor_count": len(scored), "date": today.isoformat()}


@router.post("/email-alerts")
def send_email_alerts():
    """Send expiry/breach alert emails for all active CRITICAL and HIGH alerts."""
    store = _get_store()
    entries = list(store.values())
    today = date.today()

    all_alerts = [a for e in entries for a in e.get("alerts", [])
                  if a.severity in ("CRITICAL", "HIGH")]

    if not all_alerts:
        raise HTTPException(status_code=422, detail="No CRITICAL or HIGH alerts to send.")

    try:
        from monitoring.emailer import get_emailer
        get_emailer().send_expiry_alerts(all_alerts, today)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Alert send failed: {exc}") from exc

    try:
        from api.main import app
        app.state.audit_logger.log_event(
            actor="system",
            action="alert_email_sent",
            resource_type="report",
            resource_id="expiry_alerts",
            old_state={},
            new_state={"alert_count": len(all_alerts), "date": today.isoformat()},
            reason="manual send via dashboard",
        )
    except Exception:
        pass

    return {"status": "sent", "alert_count": len(all_alerts), "date": today.isoformat()}


@router.get("/csv")
def export_csv():
    """Export all scored vendors as a flat CSV download."""
    store = _get_store()
    entries = list(store.values())
    entries.sort(key=lambda e: e["scored"].risk_score, reverse=True)

    output = io.StringIO()
    fieldnames = [
        "vendor_id", "name", "category",
        "risk_score", "risk_level", "anomaly_type", "severity",
        "contract_end", "financial_rating",
        "data_sensitivity", "access_type",
        "soc2_type2", "soc2_expiry", "iso27001", "gdpr_dpa",
        "under_investigation", "handles_eu_data",
        "breach_count",
        "top_risk_factor", "recommendation",
        "alert_count",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    today = date.today()
    for e in entries:
        v: Vendor       = e["vendor"]
        sv: ScoredVendor = e["scored"]
        alerts = e.get("alerts", [])
        writer.writerow({
            "vendor_id": v.vendor_id,
            "name": v.name,
            "category": v.category,
            "risk_score": sv.risk_score,
            "risk_level": sv.risk_level.value,
            "anomaly_type": sv.anomaly_type.value,
            "severity": sv.severity.value,
            "contract_end": v.contract_end.isoformat(),
            "financial_rating": v.financial_rating,
            "data_sensitivity": v.data_access.data_sensitivity.value,
            "access_type": v.data_access.access_type.value,
            "soc2_type2": v.compliance.soc2_type2,
            "soc2_expiry": v.compliance.soc2_expiry.isoformat() if v.compliance.soc2_expiry else "",
            "iso27001": v.compliance.iso27001,
            "gdpr_dpa": v.compliance.gdpr_dpa,
            "under_investigation": v.under_investigation,
            "handles_eu_data": v.handles_eu_data,
            "breach_count": len(v.breach_history),
            "top_risk_factor": sv.risk_factors[0] if sv.risk_factors else "",
            "recommendation": sv.recommendation,
            "alert_count": len(alerts),
        })

    output.seek(0)
    filename = f"vendor_risk_report_{today.isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/compliance-export")
def compliance_export(format: str = "json"):
    """
    Export vendor compliance summary + per-vendor compliance rows.
    ?format=json (default) or ?format=csv
    """
    store = _get_store()
    entries = list(store.values())
    today = date.today()

    vendors  = [e["vendor"]  for e in entries]
    scored   = [e["scored"]  for e in entries]

    summary = build_compliance_summary(vendors, scored, today)
    rows    = [format_vendor_for_compliance_export(v, sv) for v, sv in zip(vendors, scored)]
    rows.sort(key=lambda r: r["risk_score"], reverse=True)

    if format.lower() == "csv":
        csv_str = _build_compliance_csv(rows)
        filename = f"compliance_export_{today.isoformat()}.csv"
        return Response(
            content=csv_str,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {
        "generated_at": today.isoformat(),
        "summary": summary.model_dump(mode="json"),
        "vendors": rows,
    }


@router.get("/pdf")
def export_pdf():
    """Generate and download a PDF portfolio report using fpdf2 (pure Python, no system deps)."""
    from fpdf import FPDF

    store = _get_store()
    entries = list(store.values())
    today = date.today()
    total = len(entries)

    level_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for e in entries:
        level_counts[e["scored"].risk_level.value] += 1

    vendors = [e["vendor"] for e in entries]
    n_soc2 = sum(1 for v in vendors if v.compliance.soc2_type2
                 and (not v.compliance.soc2_expiry or v.compliance.soc2_expiry >= today))
    n_iso  = sum(1 for v in vendors if v.compliance.iso27001)
    n_gdpr = sum(1 for v in vendors if not v.handles_eu_data or v.compliance.gdpr_dpa)

    red_flags = sorted(
        [e for e in entries if e["scored"].risk_level.value in ("CRITICAL", "HIGH")],
        key=lambda e: e["scored"].risk_score, reverse=True,
    )

    LEVEL_COLOR = {
        "CRITICAL": (197, 48, 48),
        "HIGH":     (192, 86, 33),
        "MEDIUM":   (151, 90, 22),
        "LOW":      (39, 103, 73),
    }

    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    # ── Title ─────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(26, 32, 44)
    pdf.cell(0, 10, "Vendor Risk Portfolio Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(113, 128, 150)
    pdf.cell(0, 6, f"Generated: {today.isoformat()}    |    {total} vendors tracked",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_draw_color(226, 232, 240)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(6)

    # ── Risk Distribution ─────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(26, 32, 44)
    pdf.cell(0, 8, "Risk Distribution", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    box_w = 42
    for level, count in level_counts.items():
        r, g, b = LEVEL_COLOR[level]
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 20)
        pdf.cell(box_w, 12, str(count), align="C", fill=True)
        pdf.set_font("Helvetica", "", 8)
        x = pdf.get_x() - box_w
        pdf.set_xy(x, pdf.get_y() + 12)
        pdf.set_text_color(r, g, b)
        pdf.cell(box_w, 5, level, align="C")
        pdf.set_xy(pdf.get_x(), pdf.get_y() - 12)

    pdf.ln(20)

    # ── Compliance Coverage ───────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(26, 32, 44)
    pdf.cell(0, 8, "Compliance Coverage", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    comp_items = [
        (f"{round(100*n_soc2/total) if total else 0}%", "SOC 2 Type II"),
        (f"{round(100*n_iso/total) if total else 0}%",  "ISO 27001"),
        (f"{round(100*n_gdpr/total) if total else 0}%", "GDPR DPA"),
    ]
    for val, label in comp_items:
        pdf.set_fill_color(247, 248, 250)
        pdf.set_text_color(45, 55, 72)
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(55, 10, val, align="C", fill=True)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(113, 128, 150)
        x = pdf.get_x() - 55
        pdf.set_xy(x, pdf.get_y() + 10)
        pdf.cell(55, 5, label, align="C")
        pdf.set_xy(pdf.get_x(), pdf.get_y() - 10)

    pdf.ln(18)

    # ── Red-Flag Vendors table ────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(26, 32, 44)
    pdf.cell(0, 8, f"Critical & High Risk Vendors (top {min(50, len(red_flags))})",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Header row
    col_w = [22, 48, 32, 20, 18, 40]
    headers = ["ID", "Vendor", "Category", "Risk", "Score", "Top Risk Factor"]
    pdf.set_fill_color(247, 248, 250)
    pdf.set_text_color(113, 128, 150)
    pdf.set_font("Helvetica", "B", 7)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h.upper(), border="B", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for e in red_flags[:50]:
        v, sv = e["vendor"], e["scored"]
        r, g, b = LEVEL_COLOR.get(sv.risk_level.value, (51, 51, 51))
        top_factor = (sv.risk_factors[0][:45] + "...") if sv.risk_factors and len(sv.risk_factors[0]) > 45 else (sv.risk_factors[0] if sv.risk_factors else "-")

        pdf.set_text_color(45, 55, 72)
        pdf.cell(col_w[0], 6, v.vendor_id, border="B")
        name = (v.name[:26] + "...") if len(v.name) > 26 else v.name
        pdf.cell(col_w[1], 6, name, border="B")
        cat = (v.category[:20] + "...") if len(v.category) > 20 else v.category
        pdf.cell(col_w[2], 6, cat, border="B")
        pdf.set_text_color(r, g, b)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(col_w[3], 6, sv.risk_level.value, border="B")
        pdf.set_text_color(45, 55, 72)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_w[4], 6, f"{sv.risk_score:.1f}", border="B")
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(col_w[5], 6, top_factor, border="B")
        pdf.set_font("Helvetica", "", 8)
        pdf.ln()

        if pdf.get_y() > 270:  # new page before overflow
            pdf.add_page()
            pdf.set_font("Helvetica", "", 8)

    # ── Footer ────────────────────────────────────────────────────────────────
    pdf.ln(8)
    pdf.set_draw_color(226, 232, 240)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(160, 174, 192)
    pdf.cell(0, 5, f"Vendor Risk Platform  |  Confidential  |  {today.isoformat()}", align="C")

    pdf_bytes = bytes(pdf.output())
    filename = f"vendor_risk_report_{today.isoformat()}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/compliance-export", summary="Compliance export: vendor summary + audit log")
def compliance_export(format: str = "json"):
    """
    Export full vendor compliance summary plus recent audit log.
    format=json (default) returns structured JSON.
    format=csv returns a flat CSV of all vendors.
    Includes: vendor_id, name, risk_score, risk_level, soc2, iso, gdpr_dpa,
    contract_end, latest_breach_date.
    """
    store = _get_store()
    entries = list(store.values())
    today = date.today()

    rows = []
    for e in entries:
        v: Vendor = e["vendor"]
        sv: ScoredVendor = e["scored"]
        latest_breach = (
            max(b.date for b in v.breach_history).isoformat()
            if v.breach_history else None
        )
        rows.append({
            "vendor_id": v.vendor_id,
            "name": v.name,
            "category": v.category,
            "risk_score": sv.risk_score,
            "risk_level": sv.risk_level.value,
            "soc2_type2": v.compliance.soc2_type2,
            "soc2_expiry": v.compliance.soc2_expiry.isoformat() if v.compliance.soc2_expiry else None,
            "iso27001": v.compliance.iso27001,
            "gdpr_dpa": v.compliance.gdpr_dpa,
            "contract_end": v.contract_end.isoformat(),
            "under_investigation": v.under_investigation,
            "latest_breach_date": latest_breach,
        })

    # Pull audit log (last 100 events)
    audit_events: list[dict] = []
    try:
        from api.main import app
        audit_events = app.state.audit_logger.get_events(limit=100)
    except Exception:
        pass

    if format.lower() == "csv":
        output = io.StringIO()
        fieldnames = [
            "vendor_id", "name", "category", "risk_score", "risk_level",
            "soc2_type2", "soc2_expiry", "iso27001", "gdpr_dpa",
            "contract_end", "under_investigation", "latest_breach_date",
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        output.seek(0)
        filename = f"compliance_export_{today.isoformat()}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # Log this export event
    try:
        from api.main import app
        app.state.audit_logger.log_event(
            actor="system",
            action="compliance_export",
            resource_type="report",
            resource_id="compliance",
            old_state={},
            new_state={"vendor_count": len(rows), "format": format},
            reason="compliance export request",
        )
    except Exception:
        pass

    return {
        "generated_at": today.isoformat(),
        "total_vendors": len(rows),
        "vendors": rows,
        "audit_log": audit_events,
    }


@router.get("/bulk-export", summary="Bulk XLSX export of all vendors")
def bulk_export_xlsx():
    """Stream an XLSX file with all scored vendors.
    Includes compliance summary stats on a second sheet.
    Uses openpyxl.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response(
            content=b"openpyxl not installed. Run: pip install openpyxl",
            status_code=503,
            media_type="text/plain",
        )

    store = _get_store()
    entries = sorted(
        store.values(), key=lambda e: e["scored"].risk_score, reverse=True
    )
    today = date.today()

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Vendors ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Vendor Risk"

    headers = [
        "Vendor ID", "Name", "Category", "Risk Score", "Risk Level",
        "Anomaly Type", "SOC2", "ISO27001", "GDPR DPA",
        "Contract End", "Financial Rating", "Under Investigation",
        "Breach Count", "Top Risk Factor",
    ]
    LEVEL_FILL = {
        "CRITICAL": "C53030",
        "HIGH":     "C05621",
        "MEDIUM":   "975A16",
        "LOW":      "276749",
    }

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3748")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, e in enumerate(entries, 2):
        v: Vendor = e["vendor"]
        sv: ScoredVendor = e["scored"]
        level = sv.risk_level.value
        row_data = [
            v.vendor_id, v.name, v.category,
            sv.risk_score, level, sv.anomaly_type.value,
            v.compliance.soc2_type2, v.compliance.iso27001, v.compliance.gdpr_dpa,
            v.contract_end.isoformat(), v.financial_rating, v.under_investigation,
            len(v.breach_history),
            sv.risk_factors[0][:80] if sv.risk_factors else "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 5 and level in LEVEL_FILL:  # Risk Level column
                cell.fill = PatternFill("solid", fgColor=LEVEL_FILL[level])
                cell.font = Font(color="FFFFFF", bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # ── Sheet 2: Compliance Summary ───────────────────────────────────────────
    ws2 = wb.create_sheet("Compliance Summary")
    vendors = [e["vendor"] for e in entries]
    total = len(vendors)

    n_soc2 = sum(1 for v in vendors if v.compliance.soc2_type2
                 and (not v.compliance.soc2_expiry or v.compliance.soc2_expiry >= today))
    n_iso  = sum(1 for v in vendors if v.compliance.iso27001)
    n_gdpr = sum(1 for v in vendors if not v.handles_eu_data or v.compliance.gdpr_dpa)
    level_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for e in entries:
        level_counts[e["scored"].risk_level.value] += 1

    summary_rows = [
        ("Report Date", today.isoformat()),
        ("Total Vendors", total),
        ("", ""),
        ("Risk Distribution", ""),
        ("  CRITICAL", level_counts["CRITICAL"]),
        ("  HIGH", level_counts["HIGH"]),
        ("  MEDIUM", level_counts["MEDIUM"]),
        ("  LOW", level_counts["LOW"]),
        ("", ""),
        ("Compliance Coverage", ""),
        ("  SOC 2 Type II", f"{round(100*n_soc2/total) if total else 0}%"),
        ("  ISO 27001", f"{round(100*n_iso/total) if total else 0}%"),
        ("  GDPR DPA (where required)", f"{round(100*n_gdpr/total) if total else 0}%"),
    ]
    for r_idx, (label, value) in enumerate(summary_rows, 1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
        ws2.cell(row=r_idx, column=2, value=value)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20

    # Log the export
    try:
        from api.main import app
        app.state.audit_logger.log_event(
            actor="system",
            action="bulk_export",
            resource_type="report",
            resource_id="xlsx",
            old_state={},
            new_state={"vendor_count": total, "format": "xlsx"},
            reason="bulk XLSX export",
        )
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"vendor_risk_bulk_export_{today.isoformat()}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
